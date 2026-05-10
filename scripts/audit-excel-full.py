"""
Full Excel Import Audit Script
Extracts all data from every relevant sheet for comparison against the database.
"""
import openpyxl
import json
from datetime import datetime, date

EXCEL_PATH = "/home/ubuntu/upload/1112111.xlsm"

def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v) if v is not None else None

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

# ─── 1. HEADS (Animals) ───────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Heads (Animals)")
print("="*70)
ws = wb["Heads"]
rows = list(ws.iter_rows(values_only=True))
# Find header row
header = None
data_rows = []
for i, row in enumerate(rows):
    if row[0] == "ID" or (row[0] is not None and str(row[0]).strip().upper() in ["ID", "TAG", "ANIMAL_ID"]):
        header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(row)]
        data_rows = rows[i+1:]
        break
if header is None:
    # Try first non-empty row
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(row)]
            data_rows = rows[i+1:]
            break

print(f"Headers: {header}")
animals = []
for row in data_rows:
    if not any(c is not None for c in row):
        continue
    rec = dict(zip(header, row))
    animals.append(rec)

print(f"Total animal rows: {len(animals)}")
# Show first 5
for a in animals[:5]:
    print(f"  {a}")
print(f"  ... ({len(animals)-5} more)")

# ─── 2. RATION PLANS ─────────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Lists (Ration Plans / Feed Items)")
print("="*70)
ws = wb["Lists"]
rows = list(ws.iter_rows(values_only=True))
for i, row in enumerate(rows[:30]):
    if any(c is not None for c in row):
        print(f"  Row {i+1}: {row}")

# ─── 3. FEED STOCK ───────────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Feed_Stock")
print("="*70)
ws = wb["Feed_Stock"]
rows = list(ws.iter_rows(values_only=True))
for i, row in enumerate(rows):
    if any(c is not None for c in row):
        print(f"  Row {i+1}: {row}")

# ─── 4. OTHER EXPENSES ───────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Other_Expenses")
print("="*70)
if "Other_Expenses" in wb.sheetnames:
    ws = wb["Other_Expenses"]
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    print(f"Total non-empty rows: {len(non_empty)}")
    for i, row in enumerate(non_empty[:10]):
        print(f"  Row {i+1}: {row}")
    if len(non_empty) > 10:
        print(f"  ... ({len(non_empty)-10} more)")

# ─── 5. FEED LOG ─────────────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Feed_Log")
print("="*70)
if "Feed_Log" in wb.sheetnames:
    ws = wb["Feed_Log"]
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    print(f"Total non-empty rows: {len(non_empty)}")
    for i, row in enumerate(non_empty[:10]):
        print(f"  Row {i+1}: {row}")

# ─── 6. MONTHLY SUMMARY ──────────────────────────────────────────────────────
print("\n\n" + "="*70)
print("SHEET: Monthly_Summary")
print("="*70)
if "Monthly_Summary" in wb.sheetnames:
    ws = wb["Monthly_Summary"]
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    for i, row in enumerate(non_empty):
        print(f"  Row {i+1}: {row}")
