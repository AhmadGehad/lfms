"""
Excel Import Audit Script
Reads all sheets from the uploaded Excel file and prints structured data
for comparison against the database.
"""
import openpyxl
import json
from datetime import datetime

EXCEL_PATH = "/home/ubuntu/upload/1112111.xlsm"

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print("=== SHEETS ===")
print(wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # Skip empty sheets
    non_empty = [r for r in rows if any(c is not None for c in r)]
    if not non_empty:
        continue
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}  ({len(non_empty)} non-empty rows)")
    print(f"{'='*60}")
    # Print first 5 rows as header context
    for i, row in enumerate(non_empty[:5]):
        print(f"  Row {i+1}: {row}")
    if len(non_empty) > 5:
        print(f"  ... ({len(non_empty)-5} more rows)")
