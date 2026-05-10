"""
Full comparison: Excel vs Database
"""
import openpyxl
import json
from datetime import datetime, date

EXCEL_PATH = "/home/ubuntu/upload/1112111.xlsm"

def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() if v is not None else None

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# ─── ANIMALS ─────────────────────────────────────────────────────────────────
ws = wb["Heads"]
rows = list(ws.iter_rows(values_only=True))
header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(rows[2])]
animals = []
for row in rows[3:]:
    if not any(c is not None for c in row):
        continue
    rec = dict(zip(header, row))
    if rec.get("HeadID") is None:
        continue
    animals.append({
        "tag": str(rec["HeadID"]).strip(),
        "species": str(rec.get("Species","")).strip() or None,
        "sex": str(rec.get("Sex","")).strip() or None,
        "birthDate": fmt_date(rec.get("BirthDate")),
        "acquisitionType": str(rec.get("AcquisitionType","")).strip() or None,
        "acquisitionDate": fmt_date(rec.get("AcquisitionDate")),
        "status": str(rec.get("Status","")).strip() or None,
        "exitDate": fmt_date(rec.get("ExitDate")),
        "exitReason": str(rec.get("ExitReason","")).strip() or None,
    })

print("=== ANIMALS FROM EXCEL ===")
print(f"Total: {len(animals)}")
statuses = {}
for a in animals:
    s = a.get("status") or "None"
    statuses[s] = statuses.get(s, 0) + 1
print(f"By status/category: {statuses}")
exited = [a for a in animals if a.get("exitDate")]
print(f"Animals with exitDate ({len(exited)}):")
for a in exited:
    print(f"  {a['tag']} | exitDate={a['exitDate']} | exitReason={a['exitReason']}")

# ─── RATION PLANS ─────────────────────────────────────────────────────────────
ws = wb["Feed_Log"]
rows = list(ws.iter_rows(values_only=True))
ration_plans = []
for row in rows[2:]:
    if not row or row[0] is None:
        continue
    cat = str(row[0]).strip()
    if not cat or cat in ["Category", "Ration Plan (per head per day)"]:
        continue
    feed = str(row[1]).strip() if row[1] else None
    qty = row[2]
    if cat and feed and qty is not None:
        try:
            ration_plans.append({"category": cat, "feedItem": feed, "qty": float(qty)})
        except (ValueError, TypeError):
            pass

print(f"\n=== RATION PLANS FROM EXCEL ===")
print(f"Total: {len(ration_plans)}")
for rp in ration_plans:
    print(f"  {rp}")

# ─── FEED STOCK ───────────────────────────────────────────────────────────────
ws = wb["Feed_Stock"]
rows = list(ws.iter_rows(values_only=True))
purchases = []
stock_counts = []
for row in rows[2:]:
    if not row or row[0] is None:
        continue
    tx_date = fmt_date(row[0])
    feed_item = str(row[1]).strip() if len(row) > 1 and row[1] else None
    qty_in = row[2] if len(row) > 2 else None
    notes = str(row[4]).strip() if len(row) > 4 and row[4] else None
    count_date = fmt_date(row[17]) if len(row) > 17 and row[17] else None
    count_item = str(row[18]).strip() if len(row) > 18 and row[18] else None
    count_qty = row[19] if len(row) > 19 and row[19] is not None else None

    if tx_date and feed_item and qty_in:
        try:
            purchases.append({"date": tx_date, "feedItem": feed_item, "qty": float(qty_in), "notes": notes})
        except (ValueError, TypeError):
            pass
    if count_date and count_item and count_qty is not None:
        try:
            stock_counts.append({"date": count_date, "feedItem": count_item, "qty": float(count_qty)})
        except (ValueError, TypeError):
            pass

print(f"\n=== FEED PURCHASES FROM EXCEL ===")
print(f"Total: {len(purchases)}")
for p in purchases:
    print(f"  {p}")
print(f"\n=== STOCK COUNTS FROM EXCEL ===")
print(f"Total: {len(stock_counts)}")
for sc in stock_counts:
    print(f"  {sc}")

# ─── OTHER EXPENSES ───────────────────────────────────────────────────────────
ws = wb["Other_Expenses"]
rows = list(ws.iter_rows(values_only=True))
expenses = []
for row in rows[2:]:
    if not row or row[0] is None:
        continue
    exp_date = fmt_date(row[0])
    category = str(row[1]).strip() if row[1] else None
    amount = row[2]
    head_id = str(row[3]).strip() if len(row) > 3 and row[3] else None
    cat_target = str(row[4]).strip() if len(row) > 4 and row[4] else None
    vendor = str(row[5]).strip() if len(row) > 5 and row[5] else None
    notes = str(row[6]).strip() if len(row) > 6 and row[6] else None
    target_type = str(row[7]).strip() if len(row) > 7 and row[7] else None
    if exp_date and category and amount:
        try:
            expenses.append({
                "date": exp_date, "category": category, "amount": float(amount),
                "headId": head_id, "categoryTarget": cat_target,
                "vendor": vendor, "notes": notes, "targetType": target_type,
            })
        except (ValueError, TypeError):
            pass

print(f"\n=== OTHER EXPENSES FROM EXCEL ===")
print(f"Total: {len(expenses)}")
by_cat = {}
for e in expenses:
    by_cat[e["category"]] = by_cat.get(e["category"], 0) + e["amount"]
print(f"By category: {by_cat}")
print(f"Grand total: {sum(e['amount'] for e in expenses)}")
for e in expenses:
    print(f"  {e}")

output = {"animals": animals, "rationPlans": ration_plans,
          "feedPurchases": purchases, "stockCounts": stock_counts, "expenses": expenses}
with open("/tmp/excel-data.json", "w") as f:
    json.dump(output, f, indent=2, default=str)
print("\nSaved to /tmp/excel-data.json")
